from __future__ import annotations

from copy import copy
from datetime import datetime
from decimal import Decimal
import io
import logging
import os
from typing import List, Optional

from fastapi import APIRouter
from fastapi.responses import Response
from fpdf import FPDF
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

router = APIRouter(prefix="/reports", tags=["reports"])
logger = logging.getLogger(__name__)

TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "..", "templates")


class Transaction(BaseModel):
    account_code: str
    account_label: str
    department: str
    line_type: str
    amount: str
    transaction_date: str
    is_validated: bool
    label: str


class CashFlowPlan(BaseModel):
    direction: str
    amount: str
    flow_type: str
    label: str
    planned_date: str


class KpiValue(BaseModel):
    label: str
    value: str
    unit: str
    status: Optional[str]
    threshold_warn: Optional[str]
    threshold_critical: Optional[str]


class GenerateReportRequest(BaseModel):
    report_type: str
    format: str
    org_name: str
    transactions: List[Transaction]
    cash_flow_plans: List[CashFlowPlan]
    kpis: List[KpiValue]
    period_label: str


@router.post("/generate")
async def generate_report(req: GenerateReportRequest):
    logger.info(
        "Generation rapport %s format %s org: %s",
        req.report_type,
        req.format,
        req.org_name,
    )

    if req.format == "excel":
        buffer, content_type = generate_excel(req)
    else:
        buffer, content_type = generate_pdf(req)

    return Response(content=buffer, media_type=content_type)


def generate_excel(req: GenerateReportRequest):
    template_map = {
        "pl": "Compte_de_resultat.xlsx",
        "balance_sheet": "Bilan.xlsx",
        "cash_flow": "Flux_de_Tresorerie.xlsx",
    }

    template_file = template_map.get(req.report_type)

    if template_file:
        return fill_syscohada_template(template_file, req)

    return generate_simple_excel(req)


def _safe_write_cell(ws: Worksheet, row: int, column: int, value: str):
    try:
        ws.cell(row=row, column=column).value = value
    except AttributeError:
        # Ignore merged-cell write attempts and keep template structure intact.
        return


def fill_syscohada_template(template_file: str, req: GenerateReportRequest):
    template_path = os.path.join(TEMPLATES_DIR, template_file)
    wb = load_workbook(template_path)
    ws: Worksheet = wb.worksheets[0]

    today = datetime.now().strftime("%d/%m/%Y")

    for row in ws.iter_rows(min_row=2, max_row=5):
        for cell in row:
            if cell.value == "Désignation entité :":
                _safe_write_cell(ws, cell.row, cell.column + 1, req.org_name)
            if cell.value == "Excercice clos le :":
                _safe_write_cell(ws, cell.row, cell.column + 1, today)

    if req.report_type == "pl":
        _fill_compte_resultat(ws, req)
    elif req.report_type == "cash_flow":
        _fill_flux_tresorerie(ws, req)
    elif req.report_type == "balance_sheet":
        _fill_bilan(ws, req)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return buffer.getvalue(), content_type


def _fill_compte_resultat(ws: Worksheet, req: GenerateReportRequest):
    ref_mapping = {
        "TA": ["701", "7011", "7012"],
        "TB": ["702"],
        "TC": ["703", "706"],
        "TD": ["704", "705", "707", "708"],
        "RA": ["601", "6011", "6012"],
        "RC": ["602"],
        "RE": ["604", "605"],
        "RG": ["625"],
        "RH": ["621", "622", "623", "624", "626", "627", "628"],
        "RI": ["63"],
        "RJ": ["65"],
        "RK": ["641", "642", "643", "644", "645", "646"],
        "RL": ["681"],
        "RM": ["671", "672"],
        "TK": ["771", "772"],
    }

    ref_totals: dict[str, Decimal] = {}
    for ref, prefixes in ref_mapping.items():
        total = Decimal("0")
        for t in req.transactions:
            if any(t.account_code.startswith(p) for p in prefixes):
                total += Decimal(t.amount)
        ref_totals[ref] = total

    for row in ws.iter_rows():
        ref = row[0].value
        if ref and ref in ref_totals:
            row[4].value = ref_totals[ref]
            if row[4].number_format in ("General", "@", None):
                row[4].number_format = "#,##0"

    def get(ref: str) -> Decimal:
        return ref_totals.get(ref, Decimal("0"))

    aggregats: dict[str, Decimal] = {
        "XA": get("TA") - get("RA"),
        "XB": get("TA") + get("TB") + get("TC") + get("TD"),
        "XC": get("TA") + get("TB") + get("TC") + get("TD") - get("RC") - get("RE") - get("RG") - get("RH") - get("RI") - get("RJ"),
        "XD": get("TA") + get("TB") + get("TC") + get("TD") - get("RC") - get("RE") - get("RG") - get("RH") - get("RI") - get("RJ") - get("RK"),
    }
    aggregats["XE"] = aggregats["XD"] - get("RL")
    aggregats["XF"] = get("TK") - get("RM")
    aggregats["XG"] = aggregats["XE"] + aggregats["XF"]
    aggregats["XI"] = aggregats["XG"]

    for row in ws.iter_rows():
        ref = row[0].value
        if ref and ref in aggregats:
            cell = row[4]
            cell.value = aggregats[ref]
            cell.number_format = "#,##0"
            current_font = copy(cell.font) if cell.font else Font()
            current_font.bold = True
            cell.font = current_font


def _fill_flux_tresorerie(ws: Worksheet, req: GenerateReportRequest):
    plans = req.cash_flow_plans
    total_in = sum(Decimal(p.amount) for p in plans if p.direction == "IN")
    total_out = sum(Decimal(p.amount) for p in plans if p.direction == "OUT")
    net = total_in - total_out

    flux_map = {
        "ZB": total_in,
        "ZF": -total_out,
        "ZG": net,
        "ZH": net,
    }
    for row in ws.iter_rows():
        ref = row[0].value
        if ref and ref in flux_map:
            row[4].value = flux_map[ref]
            row[4].number_format = "#,##0"


def _fill_bilan(ws: Worksheet, req: GenerateReportRequest):
    revenues = sum(Decimal(t.amount) for t in req.transactions if t.line_type == "REVENUE")
    expenses = sum(Decimal(t.amount) for t in req.transactions if t.line_type == "EXPENSE")
    result_net = revenues - expenses

    for row in ws.iter_rows():
        ref = row[0].value
        if ref == "CI":
            row[10].value = result_net
            row[10].number_format = "#,##0"


def generate_simple_excel(req: GenerateReportRequest):
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Report")

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    if req.report_type == "transactions":
        ws.title = "Transactions"
        headers = [
            "Date",
            "Code",
            "Libellé",
            "Département",
            "Type",
            "Montant (FCFA)",
            "Validée",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for t in req.transactions:
            ws.append(
                [
                    t.transaction_date[:10],
                    t.account_code,
                    t.label,
                    t.department,
                    "Revenu" if t.line_type == "REVENUE" else "Charge",
                    Decimal(t.amount),
                    "Oui" if t.is_validated else "Non",
                ]
            )
            ws.cell(row=ws.max_row, column=6).number_format = '#,##0" FCFA"'

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 10

    elif req.report_type == "kpis":
        ws.title = "KPIs"
        headers = ["Indicateur", "Valeur", "Unité", "Statut", "Seuil Alerte", "Seuil Critique"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for k in req.kpis:
            ws.append(
                [
                    k.label,
                    Decimal(k.value),
                    k.unit,
                    k.status or "N/A",
                    k.threshold_warn or "-",
                    k.threshold_critical or "-",
                ]
            )

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def generate_pdf(req: GenerateReportRequest):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    terra = (196, 98, 45)
    kola = (45, 106, 79)
    gold = (184, 150, 62)
    ink = (26, 26, 46)
    text_lo = (153, 144, 168)

    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(*terra)
    pdf.cell(0, 12, "JUPITER_PLAN", ln=True)

    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(*ink)
    pdf.cell(0, 8, req.org_name, ln=True)

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(90, 85, 112)
    pdf.cell(0, 6, f"Genere le {datetime.now().strftime('%d/%m/%Y')} - {req.period_label}", ln=True)

    y_line = pdf.get_y() + 2
    pdf.line(10, y_line, 200, y_line)
    pdf.ln(6)

    titles = {
        "pl": "COMPTE DE RESULTAT",
        "balance_sheet": "BILAN COMPTABLE",
        "cash_flow": "PLAN DE TRESORERIE",
        "budget_variance": "BUDGET VS REEL",
        "transactions": "JOURNAL DES TRANSACTIONS",
        "kpis": "TABLEAU DE BORD KPIs",
    }
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*ink)
    pdf.cell(0, 10, titles.get(req.report_type, "RAPPORT"), ln=True)
    pdf.ln(4)

    if req.report_type == "pl":
        revenues = [t for t in req.transactions if t.line_type == "REVENUE"]
        expenses = [t for t in req.transactions if t.line_type == "EXPENSE"]

        total_rev = sum(Decimal(t.amount) for t in revenues)
        total_exp = sum(Decimal(t.amount) for t in expenses)
        ebitda = total_rev - total_exp

        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(*kola)
        pdf.cell(0, 8, "REVENUS", ln=True)

        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*ink)
        for t in revenues:
            pdf.cell(20, 6, t.account_code)
            pdf.cell(110, 6, t.account_label[:50])
            pdf.cell(50, 6, f"{float(Decimal(t.amount)):,.0f} FCFA", align="R")
            pdf.ln()

        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(*kola)
        pdf.cell(130, 7, "TOTAL REVENUS")
        pdf.cell(50, 7, f"{float(total_rev):,.0f} FCFA", align="R")
        pdf.ln(10)

        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(*terra)
        pdf.cell(0, 8, "CHARGES", ln=True)

        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*ink)
        for t in expenses:
            pdf.cell(20, 6, t.account_code)
            pdf.cell(110, 6, t.account_label[:50])
            pdf.cell(50, 6, f"{float(Decimal(t.amount)):,.0f} FCFA", align="R")
            pdf.ln()

        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(*terra)
        pdf.cell(130, 7, "TOTAL CHARGES")
        pdf.cell(50, 7, f"{float(total_exp):,.0f} FCFA", align="R")
        pdf.ln(8)

        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(*(kola if ebitda >= 0 else terra))
        pdf.cell(130, 9, "EBITDA")
        pdf.cell(50, 9, f"{float(ebitda):,.0f} FCFA", align="R")
        pdf.ln()

    elif req.report_type == "transactions":
        pdf.set_fill_color(*ink)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(28, 7, "Date", fill=True)
        pdf.cell(22, 7, "Code", fill=True)
        pdf.cell(70, 7, "Libelle", fill=True)
        pdf.cell(28, 7, "Type", fill=True)
        pdf.cell(42, 7, "Montant", fill=True, align="R")
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        for i, t in enumerate(req.transactions):
            bg = (255, 255, 255) if i % 2 == 0 else (244, 241, 236)
            pdf.set_fill_color(*bg)
            pdf.set_text_color(*ink)
            pdf.cell(28, 6, t.transaction_date[:10], fill=True)
            pdf.cell(22, 6, t.account_code, fill=True)
            pdf.cell(70, 6, t.label[:35], fill=True)
            pdf.set_text_color(*(kola if t.line_type == "REVENUE" else terra))
            pdf.cell(28, 6, "Revenu" if t.line_type == "REVENUE" else "Charge", fill=True)
            pdf.set_text_color(*ink)
            pdf.cell(42, 6, f"{float(Decimal(t.amount)):,.0f}", fill=True, align="R")
            pdf.ln()

    elif req.report_type == "kpis":
        for k in req.kpis:
            color = terra if k.status == "CRITICAL" else gold if k.status == "WARN" else kola
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(*ink)
            pdf.cell(130, 9, k.label)
            pdf.set_text_color(*color)
            val = float(Decimal(k.value))
            if k.unit == "%":
                display = f"{val:.2f}%"
            elif k.unit == "semaines":
                display = f"{val:.0f} {k.unit}"
            else:
                display = f"{val:,.0f} FCFA"
            pdf.cell(60, 9, display, align="R")
            pdf.ln(11)

    pdf.set_y(-20)
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(*text_lo)
    pdf.cell(
        0,
        5,
        f"Jupiter_Plan - Confidentiel - {req.org_name} - {datetime.now().strftime('%d/%m/%Y')}",
        align="C",
    )

    pdf_bytes = bytes(pdf.output(dest="S"))
    return pdf_bytes, "application/pdf"
